#!/usr/bin/env python3
import sys
import json
from openpyxl import load_workbook

def main():
    if len(sys.argv) < 2:
        print(json.dumps({"error": "No file path provided"}))
        sys.exit(1)

    file_path = sys.argv[1]

    try:
        wb = load_workbook(filename=file_path, data_only=True)
        ws = wb.active
        
        # Get max row and col
        max_row = ws.max_row
        max_col = ws.max_column
        
        if max_row < 2:
            print(json.dumps([]))
            return
            
        # Extract headers
        headers = []
        for col_idx in range(1, max_col + 1):
            val = ws.cell(row=1, column=col_idx).value
            headers.append(str(val).strip() if val is not None else f"Column_{col_idx}")
            
        data = []
        # Extract rows
        for r_idx in range(2, max_row + 1):
            row_data = {}
            is_empty = True
            for c_idx in range(1, max_col + 1):
                val = ws.cell(row=r_idx, column=c_idx).value
                # Convert None to empty string
                clean_val = str(val).strip() if val is not None else ""
                row_data[headers[c_idx - 1]] = clean_val
                if clean_val != "":
                    is_empty = False
            
            # Skip completely empty rows
            if not is_empty:
                data.append(row_data)
                
        print(json.dumps(data))
        
    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)

if __name__ == "__main__":
    main()

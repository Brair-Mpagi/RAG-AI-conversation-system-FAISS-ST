import sys
import argparse
from openpyxl import Workbook

def parse_markdown_table(filepath):
    data = []
    with open(filepath, 'r') as f:
        lines = f.readlines()
        
    for line in lines:
        line = line.strip()
        if not line or not line.startswith('|'):
            continue
        parts = [p.strip() for p in line.split('|')[1:-1]]
        if len(parts) >= 3 and parts[0].lower() != 'names':
            data.append({
                'name': parts[0],
                'position': parts[1],
                'qualification': parts[2]
            })
    return data

def generate_excel(data, output_path, parent_code):
    wb = Workbook()
    ws = wb.active
    ws.title = "Entity Import"
    
    headers = [
        'Entity Name', 'Short Name', 'Entity Type Code', 'Entity Code', 
        'Parent Code', 'Description', 'Is Active',
        'title', 'position', 'email', 'phone', 
        'specialization', 'qualifications', 'publications'
    ]
    
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
        
    for i, row_data in enumerate(data, 2):
        name = row_data['name']
        position = row_data['position']
        qualification = row_data['qualification']
        
        # generate entity code: STAFF_FIRSTNAME_LASTNAME
        name_parts = name.upper().split()
        entity_code = "STAFF_" + "_".join(name_parts)
        entity_code = "".join(c for c in entity_code if c.isalnum() or c == '_')
        
        row_values = [
            name,                      # Entity Name
            "",                        # Short Name
            "staff",                   # Entity Type Code
            entity_code,               # Entity Code
            parent_code,               # Parent Code
            f"{position} in the department", # Description
            "1",                       # Is Active
            "",                        # title
            position,                  # position
            "",                        # email
            "",                        # phone
            "",                        # specialization
            qualification,             # qualifications
            ""                         # publications
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            ws.cell(row=i, column=col_idx, value=val)
            
    wb.save(output_path)
    print(f"Saved {len(data)} staff entities to {output_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Convert staff markdown table to entity import Excel")
    parser.add_argument("input_file", help="Path to markdown text file")
    parser.add_argument("--output", default="staff_import.xlsx", help="Output Excel file path")
    parser.add_argument("--parent", default="", help="Parent Entity Code (e.g., DEPT_CS)")
    args = parser.parse_args()
    
    data = parse_markdown_table(args.input_file)
    if not data:
        print("No valid table data found.")
    else:
        generate_excel(data, args.output, args.parent)

#!/usr/bin/env python3
import sys
import argparse
import json
import mysql.connector
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

def main():
    parser = argparse.ArgumentParser(description="Generate Excel template for bulk import")
    parser.add_argument("--type", type=str, default="", help="Specific entity type to generate for")
    parser.add_argument("--output", type=str, required=True, help="Output file path")
    args = parser.parse_args()

    # DB Connection (matches PHP config)
    try:
        conn = mysql.connector.connect(
            host="localhost",
            user="root",
            password="",
            database="campus_ai_db",
            unix_socket="/opt/lampp/var/mysql/mysql.sock"
        )
        cursor = conn.cursor(dictionary=True)
    except Exception as e:
        print(f"Error connecting to database: {e}")
        sys.exit(1)

    # 1. Fetch valid Entity Types
    cursor.execute("SELECT type_name, field_schema FROM entity_types")
    entity_types = cursor.fetchall()
    valid_types = [t['type_name'] for t in entity_types]

    # 2. Fetch existing Entity Codes for Parent Code dropdown
    cursor.execute("SELECT entity_code FROM university_entities WHERE entity_code IS NOT NULL AND entity_code != ''")
    existing_entities = [row['entity_code'] for row in cursor.fetchall()]
    # Limit to 500 for dropdown to avoid Excel limits if it gets huge, but Excel supports up to 32767
    existing_entities = existing_entities[:1000] 

    # 3. Determine headers based on type
    base_headers = ['Entity Name', 'Short Name', 'Entity Type Code', 'Entity Code', 'Parent Code', 'Description', 'Is Active']
    
    schema_fields = []
    select_fields = {} # name -> options

    if args.type:
        # Find schema
        for t in entity_types:
            if t['type_name'] == args.type and t['field_schema']:
                schema = json.loads(t['field_schema'])
                if 'fields' in schema:
                    for f in schema['fields']:
                        schema_fields.append(f['name'])
                        if f.get('type') == 'select' and 'options' in f:
                            select_fields[f['name']] = f['options']
                break

    headers = base_headers + schema_fields

    # Create Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Entity Import"

    # Write Headers
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write Example Row
    example_row = ["Example " + args.type.capitalize() if args.type else "Example Entity", 
                   "EX", 
                   args.type if args.type else "faculty", 
                   "EX_CODE", 
                   "", 
                   "Example description", 
                   "1"]
    
    # Pad example row
    while len(example_row) < len(headers):
        example_row.append("")
    
    for col_idx, val in enumerate(example_row, 1):
        ws.cell(row=2, column=col_idx, value=val)

    # 4. Add Data Validation (Dropdowns)
    # Excel lists in DV must be comma-separated strings < 255 characters, or formulas referencing cells.
    # To be safe with large entity lists, we will create a hidden "Lists" sheet and reference it!
    
    lists_ws = wb.create_sheet(title="Lists")
    lists_ws.sheet_state = 'hidden'

    # Write parent codes to Lists A
    lists_ws.cell(row=1, column=1, value="Parent Codes")
    for r_idx, code in enumerate(existing_entities, 2):
        lists_ws.cell(row=r_idx, column=1, value=code)
    
    parent_range = f"Lists!$A$2:$A${len(existing_entities)+1}" if existing_entities else None

    # Write entity types to Lists B
    lists_ws.cell(row=1, column=2, value="Types")
    for r_idx, t in enumerate(valid_types, 2):
        lists_ws.cell(row=r_idx, column=2, value=t)
    type_range = f"Lists!$B$2:$B${len(valid_types)+1}" if valid_types else None

    # Write custom select fields to Lists C, D, etc.
    list_col_idx = 3
    select_ranges = {}
    for fname, options in select_fields.items():
        lists_ws.cell(row=1, column=list_col_idx, value=fname)
        for r_idx, opt in enumerate(options, 2):
            lists_ws.cell(row=r_idx, column=list_col_idx, value=opt)
        select_ranges[fname] = f"Lists!${get_column_letter(list_col_idx)}$2:${get_column_letter(list_col_idx)}${len(options)+1}"
        list_col_idx += 1

    # Apply validations to columns (apply up to row 1000)
    MAX_ROWS = 1000

    def add_validation(col_name, formula):
        col_letter = get_column_letter(headers.index(col_name) + 1)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = 'Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{MAX_ROWS}")

    if type_range:
        add_validation('Entity Type Code', type_range)
    
    if parent_range:
        add_validation('Parent Code', parent_range)
        
    for fname, formula in select_ranges.items():
        add_validation(fname, formula)

    # Apply Is Active validation directly
    is_active_letter = get_column_letter(headers.index('Is Active') + 1)
    dv_active = DataValidation(type="list", formula1='"1,0"', allow_blank=True)
    ws.add_data_validation(dv_active)
    dv_active.add(f"{is_active_letter}2:{is_active_letter}{MAX_ROWS}")

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 40) # cap at 40

    # Save
    wb.save(args.output)
    print(args.output)

if __name__ == "__main__":
    main()
